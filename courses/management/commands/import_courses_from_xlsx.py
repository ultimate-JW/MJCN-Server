"""강의시간표 엑셀(.xlsx) → Course/CourseOffering/CourseSchedule 시딩 (#36).

컬럼 15개 (spec 5.1 정의):
    학년 / 교과목명 / 과목코드 / 학과코드 / 과목번호 / 학점 / 시간 /
    담당교수 / 강좌번호 / 제한인원 / 요일 / 시작시간 / 종료시간 / 강의실 / 비고

파일명에서 연도/학기 추출 — 2026_1_컴퓨터공학전공.xlsx → year=2026, semester=1
파일명 마지막 토큰으로 college/department/major/category 기본값 매핑.
매핑 없거나 다른 값 쓰고 싶으면 --college 등 인자로 override.

사용 예:
    python manage.py import_courses_from_xlsx data/2026_1_컴퓨터공학전공.xlsx
    python manage.py import_courses_from_xlsx data/2026_1_교양.xlsx --category 교양선택
"""

import csv
import re
from datetime import time
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from courses.category_map import classify_common_area, classify_core_area, classify_liberal_subtype
from courses.models import Course, CourseOffering, CourseSchedule

# CourseSchedule.day_of_week 가능 값. 토/일 강의는 현재 모델 미지원이라 skip.
VALID_DAYS = {'월', '화', '수', '목', '금'}


# 기대 컬럼 — 헤더가 정확히 이 순서로 와야 함 (spec 5.1)
EXPECTED_COLUMNS = [
    '학년', '교과목명', '과목코드', '학과코드', '과목번호', '학점', '시간',
    '담당교수', '강좌번호', '제한인원', '요일', '시작시간', '종료시간', '강의실', '비고',
]

# 파일명 마지막 토큰 → (college, department, major, category) 기본 매핑
# 없는 토큰이면 --college / --category 인자로 보강해야 함
FILE_NAME_MAPPING = {
    '컴퓨터공학전공': ('반도체·ICT대학', '컴퓨터정보통신공학부', '컴퓨터공학전공', '전공선택'),
    '컴퓨터정보통신공학부': ('반도체·ICT대학', '컴퓨터정보통신공학부', None, '전공선택'),
    '반도체ICT대학': ('반도체·ICT대학', None, None, '전공선택'),
    # 교양 파일은 row별로 liberal_subtype 따라 category가 다름. 기본값은 '일반교양' fallback (#47 Phase 3)
    '교양': ('교양', None, None, '일반교양'),
}

# 교양 파일에서 liberal_subtype 분류 결과를 category로 사용할 영역 (#47 Phase 3)
LIBERAL_SUBTYPE_TO_CATEGORY = {
    '공통교양': '공통교양',
    '핵심교양': '핵심교양',
    '학문기초교양': '학문기초교양',
    '일반교양': '일반교양',
}

# 학과별 전공필수 과목/prefix 상수는 services 등과 공유하기 위해 별도 모듈로 분리.
from courses.required_courses import MAJOR_REQUIRED_BY_MAJOR, MAJOR_DEPT_PREFIXES

FILE_NAME_PATTERN = re.compile(r'(\d{4})_(\d)_(.+)\.xlsx$')


def parse_filename(path):
    name = Path(path).name
    m = FILE_NAME_PATTERN.match(name)
    if not m:
        return None, None, None, None
    year = int(m.group(1))
    semester = int(m.group(2))
    origin = m.group(3)  # 파일명 마지막 토큰 (예: '컴퓨터공학전공')
    return year, semester, origin, FILE_NAME_MAPPING.get(origin)


def parse_year_open(value):
    """학년 셀 → 정수. '전학년' 등 학년 무관 표현은 0."""
    s = str(value or '').strip()
    if s == '전학년':
        return 0  # 학년 무관 — 추천 알고리즘에서 별도 처리 필요 (#36 후속)
    try:
        return int(s)
    except ValueError:
        m = re.match(r'(\d+)', s)
        if m:
            return int(m.group(1))
        return 0


def to_time(value):
    """엑셀 셀 값(time / 'HH:MM' / 'HH:MM:SS')을 datetime.time으로 정규화."""
    if value is None or value == '':
        return None
    if isinstance(value, time):
        return value
    parts = str(value).strip().split(':')
    if len(parts) >= 2:
        return time(int(parts[0]), int(parts[1]))
    raise ValueError(f"시간 파싱 실패: {value!r}")


class Command(BaseCommand):
    help = '강의시간표 엑셀(.xlsx) → Course/CourseOffering/CourseSchedule 시딩 (#36)'

    def add_arguments(self, parser):
        parser.add_argument('path', help='엑셀 파일 경로')
        parser.add_argument('--year', type=int, help='연도 (없으면 파일명에서 추출)')
        parser.add_argument('--semester', type=int, help='학기 1/2/3/4 (없으면 파일명에서 추출)')
        parser.add_argument('--college', help='대학 (파일명 매핑 없으면 필요)')
        parser.add_argument('--department', help='학부')
        parser.add_argument('--major', help='전공')
        parser.add_argument('--category', help='이수구분 (전공필수/전공선택/공통교양/핵심교양/학문기초교양/일반교양/자유선택)')
        parser.add_argument(
            '--dump-csv', action='store_true',
            help='엑셀과 같은 폴더에 .csv도 함께 dump (VSCode에서 미리보기/diff용)',
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts['path']
        f_year, f_semester, origin, mapping = parse_filename(path)
        # 인자가 있으면 인자 우선, 없으면 파일명 추정값 (fallback)
        year = opts['year'] or f_year
        semester = opts['semester'] or f_semester
        if not year or not semester:
            raise CommandError(
                "연도/학기 추출 실패. 파일명 패턴(YYYY_S_*.xlsx) 또는 --year/--semester 필요"
            )

        default_college = default_dept = default_major = default_category = None
        if mapping:
            default_college, default_dept, default_major, default_category = mapping
        college = opts['college'] or default_college
        department = opts['department'] or default_dept
        major = opts['major'] or default_major
        category = opts['category'] or default_category

        if not college or not category:
            raise CommandError(
                f"college/category 필요. 파일명 토큰({origin!r}) 매핑 없으면 "
                "--college, --category 지정"
            )

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if header != EXPECTED_COLUMNS:
            raise CommandError(
                f"컬럼 불일치.\n  기대: {EXPECTED_COLUMNS}\n  실제: {header}"
            )

        # --dump-csv: 같은 폴더에 .csv 형태로도 dump (VSCode에서 바로 열림 + git diff용)
        # utf-8-sig: Excel에서도 한글 안 깨지게 BOM 포함
        csv_dump_path = None
        if opts['dump_csv']:
            csv_dump_path = Path(path).with_suffix('.csv')
            with open(csv_dump_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(EXPECTED_COLUMNS)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    writer.writerow(['' if c is None else str(c) for c in row])

        n_course_new = n_offering_new = n_schedule_new = 0
        n_rows = 0
        n_schedule_skipped = 0  # 시간/요일 누락된 행 (사이버·계약 강의 등)
        unmapped_liberal = []  # 교양 카테고리인데 4종 분류가 안 잡힌 (학과코드, 교과목명) 모음 — stdout 경고용
        unmapped_core = []     # 핵심교양인데 4영역 매핑이 안 잡힌 (학과코드, 교과목명) 모음 — category_map.CORE_AREA_BY_NAME 보강 시그널

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 학년 비어있는 행 skip (엑셀 trailing 빈줄 등)
                continue
            (year_open, name, course_code, dept_code, _course_no, credits, _hours,
             professor, section_no, capacity, day, start, end, room, note) = row
            n_rows += 1

            # 교양 4종(공통/핵심/학문기초/일반) 분류 — 학과코드 prefix + 교과목명으로 결정 (category_map).
            # 교양 파일에서만 의미 있고 전공/군사는 null 반환.
            liberal_subtype = classify_liberal_subtype(dept_code, name)

            # 교양 파일이면 liberal_subtype을 category로 박음 (#47 Phase 3 — 학칙 7분류)
            # liberal_subtype 매핑 실패 시 fallback default_category(='일반교양')로 둠 — WARN으로 표면화
            row_category = category
            if origin == '교양':
                if liberal_subtype in LIBERAL_SUBTYPE_TO_CATEGORY:
                    row_category = LIBERAL_SUBTYPE_TO_CATEGORY[liberal_subtype]
            else:
                # 전공 파일에서 학칙 §5.1 전공필수 8과목 매칭 시 category 보강 (#47)
                # 학과코드 prefix가 해당 전공 인정 범위(컴공/컴정/반아)일 때만 전공필수로 정정.
                # 매칭 실패는 정상(대다수 전공선택)이라 WARN 없음.
                prefix = (dept_code or '').strip()
                for m, names in MAJOR_REQUIRED_BY_MAJOR.items():
                    if name in names and prefix in MAJOR_DEPT_PREFIXES.get(m, ()):
                        row_category = '전공필수'
                        break

            # 핵심교양 4영역 — 핵심교양 행만 추가 분류 (#47 Phase 2). 그 외는 None.
            # 교양 영역 — core_area 필드 한 곳에 박음. liberal_subtype 따라 다른 매핑 dict 사용.
            # 핵심교양: §4.3.1~4.3.4 (역사·철학/사회·공동체/문화·예술/과학기술·정보)
            # 공통교양: §4.2.1~4.2.4 (기독교/사고와 표현/언어/진로와 디지털리터러시)
            if liberal_subtype == '핵심교양':
                core_area = classify_core_area(name)
                if core_area is None:
                    unmapped_core.append((str(dept_code or ''), str(name or '')))
            elif liberal_subtype == '공통교양':
                core_area = classify_common_area(name)
                # 공통교양 미매핑은 폐지·이수구분 변경 과목(graduation_requirements.md §9.1)일 수도 있어 WARN 안 띄움
            else:
                core_area = None

            # Course — 과목코드 unique. 같은 과목코드가 분반/요일별로 여러 행에 등장하므로 update_or_create.
            # semester_open은 권장 학기 의미라 import 학기로 일단 채움 (Course 단위로 더 정확한 값이 없음, 추후 보정)
            course, created = Course.objects.update_or_create(
                course_code=course_code,
                defaults={
                    'name': name,
                    'college': college,
                    'department': department,
                    'major': major,
                    'category': row_category,
                    'liberal_subtype': liberal_subtype,
                    'core_area': core_area,
                    'credits': int(credits) if credits else 0,
                    'year_open': parse_year_open(year_open),
                    'semester_open': int(semester),
                    'professor': professor or '',  # 분반 여러 개면 마지막 행 값으로 덮음 (Course.professor는 호환용 레거시)
                },
            )
            if created:
                n_course_new += 1

            # CourseOffering — (year, semester, section_no) unique (강좌번호는 학기 안에서 유일)
            offering, created = CourseOffering.objects.update_or_create(
                year=year, semester=semester, section_no=str(section_no),
                defaults={
                    'course': course,
                    'professor': professor or '',
                    'capacity': int(capacity) if capacity else None,
                    'note': note or '',
                },
            )
            if created:
                n_offering_new += 1

            # CourseSchedule — (offering, 요일) 1행. course FK도 동시 채움.
            # 사이버/계약 강의는 요일='미입력' + 시간/강의실 빈값 → Schedule 생성 skip (Offering 만 존재)
            start_t = to_time(start)
            end_t = to_time(end)
            day_clean = str(day or '').strip()
            if day_clean not in VALID_DAYS or start_t is None or end_t is None:
                n_schedule_skipped += 1
                continue

            room_clean = str(room or '').strip()
            if room_clean == '미입력':
                room_clean = ''
            _, created = CourseSchedule.objects.update_or_create(
                offering=offering,
                day_of_week=day_clean,
                defaults={
                    'course': course,
                    'start_time': start_t,
                    'end_time': end_t,
                    'building': '',
                    'room': room_clean,
                },
            )
            if created:
                n_schedule_new += 1

        msg = (
            f"[OK] {path}\n"
            f"  학기: {year}-{semester}\n"
            f"  매핑: college={college!r} dept={department!r} major={major!r} category={category!r}\n"
            f"  처리 행 수: {n_rows}\n"
            f"  신규 Course: {n_course_new} / Offering: {n_offering_new} / Schedule: {n_schedule_new}\n"
            f"  Schedule skip (시간·요일 누락): {n_schedule_skipped}"
        )
        if csv_dump_path:
            msg += f"\n  CSV dump: {csv_dump_path}"
        self.stdout.write(self.style.SUCCESS(msg))

        # 4종 매핑 안 잡힌 교양 과목 경고 — 학과코드/과목명 (중복 제거)
        if unmapped_liberal:
            unique = sorted(set(unmapped_liberal))
            self.stdout.write(self.style.WARNING(
                f"  [WARN] 교양 4종 분류 미매핑 {len(unique)}건 (category_map.py 보강 필요):"
            ))
            for dept_code, cname in unique:
                self.stdout.write(f"    - [{dept_code}] {cname}")

        # 핵심교양 4영역 매핑 안 잡힌 행 경고 — CORE_AREA_BY_NAME 보강 시그널
        if unmapped_core:
            unique = sorted(set(unmapped_core))
            self.stdout.write(self.style.WARNING(
                f"  [WARN] 핵심교양 4영역 미매핑 {len(unique)}건 (CORE_AREA_BY_NAME 보강 필요):"
            ))
            for dept_code, cname in unique:
                self.stdout.write(f"    - [{dept_code}] {cname}")
